import requests
from requests.models import PreparedRequest
from bs4 import BeautifulSoup
import openpyxl
from datetime import date
from configparser import ConfigParser
import math
        

def printOptionsAndGetMapping(config: ConfigParser):
    print()
    sections = config.sections()
    keysMapping = {}
    count = 0
    for section in sections:
        level = 0
        if section.count('-') > 1:
            level += section.count('-') - 1

        if 'КУПИТЬ' in section or 'СНЯТЬ' in section:
            level += 1
        
        print(f'{"   "*level}{section}')

        for key in config[section].keys():
            count += 1
            print(f'   {"   "*level}{count} {key.capitalize()} = {config[section][key]}')
            keysMapping[count] = {'section': section, 'key': key}
        print()
    
    return keysMapping

def containSection(sectionWords, sectionName):
    return len([word for word in sectionWords if word in sectionName]) > 0


def makeConfigRecord(config, mapping, paramId):
    sectionName = mapping[paramId]['section']
    key = mapping[paramId]['key']

    if containSection(['ШАГИ', 'ЛИФТ', 'ПЕРИОДЫ', 'РАЙОНЫ', 'РЕМОНТ', 'СОБСТВЕННИК'], sectionName):
        if config[sectionName][key] == '':
            config[sectionName][key] = 'да'
        else:
            config[sectionName][key] = ''
    else:
        value = input('Введите значение или нажмите enter для удаления: ').strip().lower()
        if value:
            if 'площади' in key:
                while ' -' in value or '- ' in value or '  ' in value or '--' in value:
                    value = value.replace(' -', '-')
                    value = value.replace('- ', '-')
                    value = value.replace('  ', ' ')
                    value = value.replace('--', '-')
            
            config[sectionName][key] = value
        else:
            config[sectionName][key] = ''
    

def editConfig(config: ConfigParser):
    mapping = printOptionsAndGetMapping(config)
    prompt = 'Введите номер параметра для редактирования или нажмите enter чтобы запустить обработку: '
    errorPrompt = 'Введено некорректное значение'
    paramId = input(prompt)
    while paramId != '':
        try:
            paramId = int(paramId)
            if paramId < 1 or paramId > max(list(mapping.keys())):
                print(errorPrompt)                
            else:
                makeConfigRecord(config, mapping, paramId)        
                mapping = printOptionsAndGetMapping(config)
                print('Изменения внесены')
            paramId = input(prompt)
        except:
            print(errorPrompt)
            paramId = input(prompt)

    with open('cache.ini', 'w', encoding="utf-8") as file:
        config.write(file)

    print('\nФайл конфигурации сохранён')

def getParams(config):
    params = {}
    for option in config:
        if option.get('extra', False):
            for extra in option['extra']:
                if extra['value']:
                    params[extra['urlParam']] = extra['value']
        if option.get('urlParam', False) and option.get('value', False):
            params[option['urlParam']] = option['value']
    
    return params

def getResponse(url, params):
    print('Произвожу запрос')

    req = PreparedRequest()
    req.prepare_url(url, params)
    count = 2
    while True:
        try:
            response = requests.get(
                url='https://proxy.scrapeops.io/v1/',
                params={
                    'api_key': 'badc2afe-6a80-44e1-a995-c3764685e50b',
                    'url': req.url.replace('%25', '%'),
                },
            )
            if response.ok:
                print(f'Ответ получен [{response.status_code}]')
                break
        except:
            print('Ошибка подключения к прокси')    
        print(f'Попытка №{count}')
        count += 1
    
    return response

def retrieveCount(result):
    return int(result[result.find(' ')+1:result.rfind(' ')].replace(' ', ''))

def getCountText(config, url):
    soup = BeautifulSoup(getResponse(url, getParams(config)).text, 'html.parser')
    summary = soup.select_one('div[data-name="SummaryHeader"]')
    return summary.text if summary != None else 'Найдено 0 объявлений'

def getMiss(half, options, optimizations, url):
    print('Проверяю цену: ' + optimizations[0]['value'])
    innerCount = retrieveCount(getCountText(options + optimizations, url))
    miss = half - innerCount
    print(f'Погрешность: {miss}\n')

    return miss

def getMissTuple(count, options, optimizations, url, isBuy):
    if count == 0:
        return [0, optimizations[0]['value']]
    if count < 200:
        print(f'Количество объявлений < 200')
        print(f'Произвожу расчёт по страницам\n')
        return pageBasedMissTupleGetter(options, optimizations, isBuy, url)
    return percentBasedMissTupleGetter(count, options, optimizations, url)

def pageBasedMissTupleGetter(options, optimizations, isBuy, url):
    roundStep = int(optimizations[1]['value'])

    params = getParams(options)
    soup = BeautifulSoup(getResponse(url, params).text, 'html.parser')
    summary = soup.select_one('div[data-name="SummaryHeader"]')

    total_count = retrieveCount(summary.text)
    target_listing = round(total_count / 2)
    listings = soup.select('article[data-name="CardComponent"]')
    count_on_page_avg = len(listings)
    target_page = math.ceil(target_listing / count_on_page_avg)

    print(f'Объявлений на странице: {count_on_page_avg}')
    print(f'Искомая цена на странице № {target_page}\n' )

    if target_page != 1:
        params['p'] = target_page
        soup = BeautifulSoup(getResponse(url, params).text, 'html.parser')
        listings = soup.select('article[data-name="CardComponent"]')

    target_listing_index = target_listing - 1
    min_index_of_total_on_page = count_on_page_avg * (target_page-1)
    to_pick_index = target_listing_index - min_index_of_total_on_page

    if isBuy:
        price_text = listings[to_pick_index].select_one('p[data-mark="PriceInfo"]').text
        price_text = price_text[0:price_text.rfind(' ')+4] + ' Круто'
    else:
        price_text = listings[to_pick_index].select_one('span[data-mark="MainPrice"]').text
    price = retrieveCount(f"Цена: {price_text}")
    return [0, roundToStep(price, roundStep)]

def percentBasedMissTupleGetter(count, options, optimizations, url):
    roundStep = int(optimizations[1]['value'])
    half = count // 2
    miss = getMiss(half, options, optimizations, url)
    history = {miss: optimizations[0]['value']}
    
    changes = 0
    toCheck = getToCheckPercentBased(count, miss, optimizations[0]['value'], roundStep, changes)
    isPositive = miss > 0

    while (toCheck != optimizations[0]['value']) and (toCheck not in list(history.values())):
        optimizations[0]['value'] = toCheck
        miss = getMiss(half, options, optimizations, url)
        history[miss] = toCheck

        if (miss > 0) != isPositive:
            isPositive = not isPositive
            changes += 1

        optimized = False
        for key in list(history.keys()):
            if (miss > 0 and key < 0) or (miss < 0 and key > 0):
                ratio = abs(miss / key)
                if ratio >= 0.9 and ratio <= 1.1:
                    toCheck = roundToStep((int(toCheck)+ int(history[key])) / 2, roundStep)
                    optimized = True
                    break
        
        if not optimized:
            toCheck = getToCheckPercentBased(count, miss, toCheck, roundStep, changes)

    min_key = min([abs(key) for key in list(history.keys())])
    return [min_key, history.get(min_key, False) or history.get(-min_key, False)]

def getToCheckPercentBased(count, miss, currentPrice, roundStep, changesCount):
    if miss == 0:
        return currentPrice

    currentPrice = int(currentPrice)
    roundStep = int(roundStep)

    percent = miss / count
    percent /= 2**changesCount


    if percent > 0:
        percent = min(0.3, percent)
        return roundToStep(currentPrice*(1+percent), roundStep)
    else:
        percent = max(-0.3, percent)
        return roundToStep(currentPrice+currentPrice*percent, roundStep)

def roundToStep(n, roundStep):
    i = int(n // roundStep)
    rem = n % roundStep
    if rem >= (roundStep // 2):
        i+= 1
    return str(roundStep*i)

def saveToCache(config: ConfigParser, sectionName, propertyName, avg):
    config[sectionName][propertyName] = avg
    with open('cache.ini', 'w', encoding="utf-8") as file:
        config.write(file)

class ExcelWorker:
    def __init__(self, fileName) -> None:
        self.workBook = openpyxl.load_workbook(f'{fileName}.xlsx')
        self.fileName = fileName

    def createRow(self, options, sheetName):
        sheet = self.workBook[sheetName]
        max_row = sheet.max_row+1
        current_column = sheet.min_column

        sheet.cell(row=max_row, column=current_column, value=date.today())
        for option in options:
            if not option.get('hide', False):
                current_column += 1
                value = option['value'] if option.get('label', None) == None else option['label']
                sheet.cell(row=max_row, column=current_column, value=value)

        self.current_column = current_column
        self.current_sheetName = sheetName

        self.save()

    def addToRow(self, count, avg):
        sheet = self.workBook[self.current_sheetName]
        max_row = sheet.max_row
        current_column = self.current_column
        sheet.cell(row=max_row, column=current_column+1, value=count)
        sheet.cell(row=max_row, column=current_column+2, value=avg)

        self.current_column = current_column+2

        self.save()

    def moveColumnsPointer(self, count):
        self.current_column = self.current_column + count

    def save(self):
        self.workBook.save(f'{self.fileName}.xlsx')

    def close(self):
        self.workBook.close()

    


